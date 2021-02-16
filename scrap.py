import datetime
import requests
import logging



from openpyxl.workbook import Workbook
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
def scraper():
    speakers=[]
    for i in range(1,16):
        link="https://www.amazon.in/s?k=top+50+speakers&page="+str(i)+"&qid=1612250144&ref=sr_pg_"+str(i)

        

        site=requests.get(link).text
        
        s=BeautifulSoup(site,'lxml')
        main=s.findAll("div",attrs={'class':'s-include-content-margin s-border-bottom s-latency-cf-section'})

        for d in main:
    
            name=d.find('span', attrs={'class':'a-size-medium a-color-base a-text-normal' }).text
    
            if d.find('span', attrs={'class':'a-price-whole'}):
                s=d.find('span', attrs={'class':'a-price-symbol'}).text
                l=d.find('span', attrs={'class':'a-price-whole'}).text
                Listed_price= str(s)+str(l)
            else:
                Listed_price="NA"
            if d.find('span', attrs={'class':'a-price a-text-price'}):
                a=d.find('span', attrs={'class':'a-price a-text-price'})
                if a.find('span', attrs={'class':'a-offscreen'}):
                    Actual_price=a.find('span', attrs={'class':'a-offscreen'}).text
                else:
                    Actual_price="NA"
            if d.find('span', attrs={'class':'a-icon-alt'}):
                rating=d.find('span', attrs={'class':'a-icon-alt'}).text[0:3]
            else:
                rating="NA"
            dt=datetime.datetime.now()
            #print(dt)
        
            speakers.append([name,Listed_price,Actual_price,rating,str(dt)])
    return speakers                

def data_to_xlsx(speakers):

    wb=Workbook()
        # grab the active worksheet
    ws = wb.active
        # Data can be assigned directly to cells
    ws['A1'].fill = PatternFill(start_color="0000CCFF", fill_type = "solid")
    ws['A1'] = 'Name'
    ws['B1'].fill = PatternFill(start_color="0000CCFF", fill_type = "solid")
    ws['B1'] = 'Listing Price'
    ws['c1'].fill = PatternFill(start_color="0000CCFF", fill_type = "solid")
    ws['C1'] = 'Actual Price'
    ws['D1'].fill = PatternFill(start_color="0000CCFF", fill_type = "solid")
    ws['D1'] = 'Ratings'
    ws['E1'].fill = PatternFill(start_color="0000CCFF", fill_type = "solid")
    ws['E1'] = 'date_time'
    for i in speakers:
        
        ws.append(i)
        
    # Save the file
    wb.save("amazon_top_speakers.xlsx")
    return 0

def log_to_xlsx(speakers):

    log_records=[]
    for i in speakers:
        log_records.append(i)
        logging.info(i)
    wb_log=Workbook()
    ws_log=wb_log.active
    ws_log['A1']='Logging_Info'
    for logs in log_records:
        ws_log.append(logs)
    # wb_log.save("Logging Report.xlsx")
result=scraper()
data_to_xlsx(result)
